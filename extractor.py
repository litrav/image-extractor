import os
import zipfile
import re  # Essencial, como você disse!
from pathlib import Path
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

# ---
# NOTA: As funções 'extrair_todas_imagens' e 'extrair_imagens_por_aba'
# e 'main' estão perfeitas. O "coração" é a 'mapear_imagens_abas'.
# Vou colar o script inteiro para garantir.
# ---

def extrair_imagens_excel(caminho_planilha):
    """
    Extrai todas as imagens de um arquivo Excel (.xlsx, .xlsm)
    """
    print(f"\n📊 Processando planilha: {caminho_planilha}")
    
    # Perguntar nome base das imagens
    nome_base = input("\n📝 Digite o nome base para as imagens (ex: 'imagem'): ").strip()
    if not nome_base:
        nome_base = "imagem"
    
    # Perguntar nome da pasta
    nome_pasta = input("📁 Digite o nome da pasta principal: ").strip()
    if not nome_pasta:
        nome_pasta = "imagens_extraidas"
    
    # Criar pasta principal
    pasta_principal = Path(nome_pasta)
    pasta_principal.mkdir(exist_ok=True)
    print(f"\n✅ Pasta criada: {pasta_principal.absolute()}")
    
    # Carregar workbook para obter nomes das abas
    try:
        wb = load_workbook(caminho_planilha, data_only=True)
        nomes_abas = wb.sheetnames
        num_abas = len(nomes_abas)
        print(f"\n📑 Encontradas {num_abas} aba(s): {', '.join(nomes_abas)}")
    except Exception as e:
        print(f"❌ Erro ao carregar planilha: {e}")
        return
    
    # Arquivos Excel (.xlsx) são arquivos ZIP
    try:
        with zipfile.ZipFile(caminho_planilha, 'r') as zip_ref:
            # Listar todos os arquivos no ZIP
            arquivos_zip = zip_ref.namelist()
            
            # Encontrar imagens na pasta xl/media/
            imagens_media_total = [f for f in arquivos_zip if f.startswith('xl/media/')]
            
            if not imagens_media_total:
                print("\n⚠️  Nenhuma imagem encontrada na planilha")
                return
            
            print(f"\n📸 Total de {len(imagens_media_total)} imagem(ns) encontrada(s)")
            
            # Mapear imagens para abas usando arquivos de relacionamento
            mapa_aba_imagens = mapear_imagens_abas(zip_ref, arquivos_zip, nomes_abas)
            
            # --- LÓGICA DE EXTRAÇÃO CENTRALIZADA ---
            
            imagens_salvas = set()
            
            # 1. Tentar salvar por aba (se houver mapa E mais de 1 aba)
            if mapa_aba_imagens and num_abas > 1:
                imagens_salvas = extrair_imagens_por_aba(zip_ref, mapa_aba_imagens, pasta_principal, nome_base, nomes_abas)
            
            # 2. Agora, processar TODAS as imagens que NÃO foram salvas
            
            # Filtrar as que ainda não foram salvas
            imagens_nao_salvas = [img for img in imagens_media_total if img not in imagens_salvas]
            
            if imagens_nao_salvas:
                if imagens_salvas: # Se já salvamos algumas, só avisar das restantes
                    print(f"\n📸 Encontradas {len(imagens_nao_salvas)} imagem(ns) não mapeada(s) (ex: cabeçalhos, rodapés, 'fantasmas')")
                    print(f"   📂 Salvando na pasta principal...")
                else: # Se não salvamos NENHUMA, é o cenário de fallback
                    print("\n⚠️  Não foi possível mapear imagens por aba (ou há apenas 1 aba).")
                    print(f"   📂 Salvando todas as {len(imagens_nao_salvas)} imagens na pasta principal...")
                
                # Chamar a função 'extrair_todas_imagens' para salvar o restante
                extrair_todas_imagens(zip_ref, imagens_nao_salvas, pasta_principal, nome_base, len(imagens_salvas))
            
            elif not imagens_salvas and not imagens_nao_salvas:
                # Caso onde não há imagens (já tratado, mas bom verificar)
                pass 
            elif not imagens_nao_salvas and imagens_salvas:
                print("\n🎉 Todas as imagens foram mapeadas e salvas com sucesso!")
            
            # --- FIM DA LÓGICA ---
            
    except Exception as e:
        print(f"❌ Erro ao processar arquivo: {e}")
        import traceback
        traceback.print_exc()


def mapear_imagens_abas(zip_ref, arquivos_zip, nomes_abas):
    """
    Tenta mapear quais imagens pertencem a quais abas
    Esta é a parte complexa que varre os XMLs de relacionamento.
    """
    mapa = {}
    
    try:
        # --- Parte 1: Links diretos (Planilha -> Imagem) ---
        rels_files = [f for f in arquivos_zip if f.startswith('xl/worksheets/_rels/') and f.endswith('.rels')]
        print(f"\n🔍 Analisando relacionamentos de {len(rels_files)} aba(s)...")
        
        for rel_file in rels_files:
            
            # Extrair número da planilha do nome do arquivo (ex: sheet1.xml.rels)
            match = re.search(r'sheet(\d+)\.xml\.rels', rel_file)
            sheet_num = match.group(1) if match else None
            
            if not sheet_num:
                continue
            
            with zip_ref.open(rel_file) as f:
                tree = ET.parse(f)
                root = tree.getroot()
                ns = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                
                imagens_aba = []
                for rel in root.findall('.//r:Relationship', ns):
                    rel_type = rel.get('Type', '')
                    target = rel.get('Target', '')
                    
                    # Verificar se é uma imagem (pode ser image, drawing, chart, etc)
                    # O target ../media/ é o mais confiável
                    if '../media/' in target or '/media/' in target:
                        if '../media/' in target:
                            img_name = target.split('../media/')[-1]
                        elif '/media/' in target:
                            img_name = target.split('/media/')[-1]
                        else:
                            continue
                        
                        img_path = f'xl/media/{img_name}'
                        if img_path not in imagens_aba:
                            imagens_aba.append(img_path)
                
                if imagens_aba:
                    sheet_idx = int(sheet_num) - 1
                    if sheet_idx < len(nomes_abas):
                        nome_aba = nomes_abas[sheet_idx]
                        mapa[nome_aba] = imagens_aba
                        print(f"   ✅ '{nome_aba}': {len(imagens_aba)} imagem(ns) [Link Direto]")
                    else:
                        mapa[f"Aba {sheet_num}"] = imagens_aba
                        print(f"   ✅ 'Aba {sheet_num}': {len(imagens_aba)} imagem(ns) [Link Direto]")
        
        # --- Parte 2: Links indiretos (Planilha -> Drawing -> Imagem) ---
        drawing_rels = [f for f in arquivos_zip if f.startswith('xl/drawings/_rels/') and f.endswith('.rels')]
        
        if drawing_rels:
            print(f"\n🔍 Analisando {len(drawing_rels)} 'drawings' (links indiretos)...")
            
            # 2a. Mapear quais abas usam quais drawings
            drawing_to_sheet = {}
            sheet_rels = [f for f in arquivos_zip if f.startswith('xl/worksheets/_rels/') and f.endswith('.rels')]
            
            for sheet_rel in sheet_rels:
                match = re.search(r'sheet(\d+)\.xml\.rels', sheet_rel)
                sheet_num = match.group(1) if match else None
                if not sheet_num:
                    continue
                
                with zip_ref.open(sheet_rel) as f:
                    tree = ET.parse(f)
                    root = tree.getroot()
                    ns = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                    
                    for rel in root.findall('.//r:Relationship', ns):
                        target = rel.get('Target', '')
                        # O tipo de relação de drawing é diferente
                        if '../drawings/drawing' in target:
                            drawing_num = re.search(r'drawing(\d+)\.xml', target).group(1)
                            drawing_to_sheet[drawing_num] = int(sheet_num)
            
            # 2b. Mapear quais drawings usam quais imagens
            for drawing_rel in drawing_rels:
                drawing_num_match = re.search(r'drawing(\d+)\.xml\.rels', drawing_rel)
                drawing_num = drawing_num_match.group(1) if drawing_num_match else None

                if not drawing_num or drawing_num not in drawing_to_sheet:
                    continue
                
                with zip_ref.open(drawing_rel) as f:
                    tree = ET.parse(f)
                    root = tree.getroot()
                    ns = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
                    
                    imagens_drawing = []
                    for rel in root.findall('.//r:Relationship', ns):
                        target = rel.get('Target', '')
                        if '../media/' in target:
                            img_name = target.split('../media/')[-1]
                            img_path = f'xl/media/{img_name}'
                            if img_path not in imagens_drawing:
                                imagens_drawing.append(img_path)
                    
                    if imagens_drawing:
                        sheet_idx = drawing_to_sheet[drawing_num] - 1
                        if sheet_idx < len(nomes_abas):
                            nome_aba = nomes_abas[sheet_idx]
                            
                            # Adicionar no mapa (sem sobrescrever os links diretos)
                            if nome_aba not in mapa:
                                mapa[nome_aba] = []
                            
                            novas_imgs = 0
                            for img in imagens_drawing:
                                if img not in mapa[nome_aba]:
                                    mapa[nome_aba].append(img)
                                    novas_imgs += 1
                            
                            if novas_imgs > 0:
                                print(f"   ✅ '{nome_aba}': +{novas_imgs} imagem(ns) [Via Drawing]")
        
        if mapa:
            print(f"\n✅ Mapeamento concluído: {len(mapa)} aba(s) com imagens")
        
        return mapa if mapa else None
        
    except Exception as e:
        print(f"⚠️  Erro ao mapear imagens por aba: {e}")
        import traceback
        traceback.print_exc()
        return None


def extrair_todas_imagens(zip_ref, lista_imagens, pasta_destino, nome_base, contador_inicial=0):
    """
    Extrai uma lista específica de imagens para a pasta principal (fallback)
    """
    contador = contador_inicial + 1
    
    for img_path in lista_imagens:
        try:
            # Ler imagem do ZIP
            img_data = zip_ref.read(img_path)
            
            # Obter extensão original
            extensao = Path(img_path).suffix
            if not extensao:
                extensao = '.png' # Fallback
            
            # Nome do arquivo
            nome_arquivo = f"{nome_base}_{contador}{extensao}"
            caminho_completo = pasta_destino / nome_arquivo
            
            # Salvar imagem
            with open(caminho_completo, 'wb') as f:
                f.write(img_data)
            
            print(f"   ✅ Salva (não mapeada): {nome_arquivo}")
            contador += 1
            
        except Exception as e:
            print(f"   ❌ Erro ao extrair {img_path}: {e}")
    
    print(f"\n{'='*50}")
    print(f"🎉 Extração (pasta principal) concluída!")
    print(f"📊 Total de imagens extraídas nesta etapa: {contador - contador_inicial - 1}")
    print(f"📁 Pasta: {pasta_destino.absolute()}")
    print(f"{'='*50}\n")


def extrair_imagens_por_aba(zip_ref, mapa_aba_imagens, pasta_principal, nome_base, nomes_abas):
    """
    Extrai imagens organizadas por aba
    Retorna: um 'set' com os caminhos das imagens que foram salvas
    """
    contador_global = 1
    total_imagens = 0
    imagens_salvas = set() # Irá rastrear o que foi salvo
    
    # Processar imagens mapeadas por aba
    for idx_aba, nome_aba in enumerate(nomes_abas, 1):
        if nome_aba not in mapa_aba_imagens:
            continue
        
        imagens_aba = mapa_aba_imagens[nome_aba]
        num_imagens = len(imagens_aba)
        print(f"\n📸 Aba '{nome_aba}': {num_imagens} imagem(ns) encontrada(s)")
        
        # Criar subpasta para a aba
        # Limpar nome da aba para ser seguro para nomes de pasta
        nome_aba_limpo = re.sub(r'[\\/*?:"<>|]', '_', nome_aba) # Substitui inválidos
        nome_aba_limpo = re.sub(r' +', ' ', nome_aba_limpo).strip() # Remove espaços extras
        
        # Prevenir nomes de pasta muito longos
        nome_aba_limpo = nome_aba_limpo[:100]
        
        pasta_aba = pasta_principal / f"aba_{idx_aba:02d}_{nome_aba_limpo}"
        pasta_aba.mkdir(exist_ok=True)
        print(f"   📂 Subpasta criada: {pasta_aba.name}")
        
        # Extrair imagens da aba
        for img_path in imagens_aba:
            # Verificar se o arquivo de imagem realmente existe no zip
            if img_path not in zip_ref.namelist():
                print(f"   ⚠️  Aviso: Mapeamento {img_path} não encontrado no ZIP.")
                continue
                
            try:
                # Ler imagem do ZIP
                img_data = zip_ref.read(img_path)
                
                # Obter extensão original
                extensao = Path(img_path).suffix
                if not extensao:
                    extensao = '.png'
                
                # Nome do arquivo
                nome_arquivo = f"{nome_base}_{contador_global:03d}{extensao}"
                caminho_completo = pasta_aba / nome_arquivo
                
                # Salvar imagem
                with open(caminho_completo, 'wb') as f:
                    f.write(img_data)
                
                print(f"   ✅ Salva (mapeada): {nome_arquivo}")
                contador_global += 1
                total_imagens += 1
                imagens_salvas.add(img_path) # Adiciona ao set de salvas
                
            except Exception as e:
                print(f"   ❌ Erro ao extrair {img_path}: {e}")
    
    print(f"\n{'='*50}")
    print(f"🎉 Extração por abas concluída!")
    print(f"📊 Total de imagens mapeadas e salvas: {total_imagens}")
    print(f"📁 Pastas criadas dentro de: {pasta_principal.absolute()}")
    print(f"{'='*50}\n")
    
    return imagens_salvas # Retorna o set de imagens salvas


def main():
    print("="*50)
    print("🖼️   EXTRATOR DE IMAGENS DE PLANILHAS EXCEL")
    print("="*50)
    
    # Solicitar caminho da planilha
    caminho = input("\n📂 Digite o caminho completo da planilha Excel: ").strip()
    
    # Remover aspas se o usuário copiar/colar o caminho
    caminho = caminho.strip('"').strip("'")
    
    # Verificar se o arquivo existe
    if not os.path.exists(caminho):
        print(f"\n❌ Erro: Arquivo não encontrado: {caminho}")
        return
    
    # Verificar extensão
    extensao = os.path.splitext(caminho)[1].lower()
    if extensao not in ['.xlsx', '.xlsm']:
        print(f"\n⚠️  Aviso: Este script funciona com arquivos .xlsx ou .xlsm")
        print(f"   Extensão detectada: {extensao}")
        if extensao == '.xls':
            print("   ℹ️  Arquivos .xls (Excel antigo) não são suportados.")
            print("   Abra o arquivo no Excel e salve como .xlsx")
            return
        continuar = input("   Deseja tentar mesmo assim? (s/n): ").lower()
        if continuar != 's':
            return
    
    # Extrair imagens
    extrair_imagens_excel(caminho)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Operação cancelada pelo usuário")
    except Exception as e:
        print(f"\n❌ Erro inesperado: {e}")
        import traceback
        traceback.print_exc()